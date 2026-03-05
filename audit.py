import pandas as pd
import json
import openai

# API 키 설정 (본인의 OpenAI API 키 입력)
openai.api_key = "sk-your-api-key-here"

def analyze_excel_structure_with_llm(sample_csv_text):
    """
    LLM을 사용하여 엑셀 샘플 데이터에서 질문/답변 컬럼 위치와 데이터 시작 행을 찾습니다.
    """
    system_prompt = """
    너는 데이터 분석 전문가야. 
    사용자가 엑셀 파일의 상위 20행을 CSV 형식으로 제공할 거야.
    이 데이터를 분석해서 '질문(Question)'과 '답변(Answer)'에 해당하는 컬럼의 인덱스(0부터 시작)와,
    실제 데이터(헤더 제외)가 시작되는 행 인덱스(0부터 시작)를 찾아줘.
    
    반드시 아래와 같은 JSON 형식으로만 대답해:
    {
        "start_row_idx": 5, 
        "question_col_idx": 1, 
        "answer_col_idx": 2
    }
    """

    response = openai.chat.completions.create(
        model="gpt-4o", # 빠르고 저렴한 gpt-4o-mini를 사용해도 좋습니다.
        response_format={ "type": "json_object" },
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"다음은 엑셀 샘플 데이터야:\n\n{sample_csv_text}"}
        ],
        temperature=0.1 # 일관된 결과를 위해 온도를 낮춤
    )
    
    # LLM이 응답한 JSON 문자열을 파이썬 딕셔너리로 변환
    result_json = response.choices[0].message.content
    return json.loads(result_json)

def extract_qna_from_excel(file_path):
    """
    엑셀 파일을 읽어 질문과 답변 데이터만 추출합니다.
    """
    print(f"[{file_path}] 파일 분석 시작...")

    # 1. 샘플링: 컬럼명이나 형식을 모르므로 헤더 없이 상위 20행만 읽어옴
    try:
        sample_df = pd.read_excel(file_path, header=None, nrows=20)
    except Exception as e:
        return f"엑셀 파일을 읽는 중 오류 발생: {e}"

    # 결측치(NaN)를 빈 문자열로 처리하여 LLM이 헷갈리지 않게 함
    sample_csv = sample_df.fillna("").to_csv(index=False, header=False)

    # 2. LLM에게 구조 파악 요청
    print("LLM에게 구조 분석 요청 중...")
    try:
        structure = analyze_excel_structure_with_llm(sample_csv)
        start_row = structure.get("start_row_idx", 0)
        q_col = structure.get("question_col_idx", 0)
        a_col = structure.get("answer_col_idx", 1)
        print(f"✅ 분석 완료! 데이터 시작 행: {start_row}, 질문 컬럼: {q_col}, 답변 컬럼: {a_col}")
    except Exception as e:
        return f"LLM 분석 중 오류 발생: {e}"

    # 3. 파악된 구조를 바탕으로 전체 데이터 정확히 추출
    # header 매개변수를 사용하여 데이터가 시작되는 바로 윗줄을 헤더로 인식하게 하거나,
    # 인덱스 기반으로 슬라이싱할 수 있습니다. 여기서는 인덱스 슬라이싱 사용.
    
    full_df = pd.read_excel(file_path, header=None)
    
    # 실제 데이터가 있는 행부터 끝까지 자르기
    data_df = full_df.iloc[start_row:].copy()
    
    # 질문과 답변 컬럼만 선택
    try:
        qna_df = data_df.iloc[:, [q_col, a_col]].copy()
    except IndexError:
        return "LLM이 잘못된 컬럼 인덱스를 반환했습니다. 엑셀 양식을 확인해주세요."

    # 컬럼명 통일 (프론트엔드로 보내기 좋게)
    qna_df.columns = ["Question", "Answer"]
    
    # 질문이나 답변이 완전히 비어있는 행은 제거 (정제 과정)
    qna_df = qna_df.dropna(how='all')

    print("🎉 데이터 추출 성공!")
    return qna_df

# --- 실행 예시 ---
if __name__ == "__main__":
    # 테스트용 엑셀 파일 경로
    # test_file = "sample_qna.xlsx" 
    # extracted_data = extract_qna_from_excel(test_file)
    # print(extracted_data.head())
    
    # 프론트엔드로 전달할 때는 JSON 형태로 변환하면 좋습니다.
    # frontend_json = extracted_data.to_dict(orient="records")
    pass



import openpyxl
import json
import openai

# openai.api_key = "sk-your-api-key"

def find_qna_coordinates(file_path, sheet_name):
    print(f"[{sheet_name}] 시트 좌표 탐색 시작...")
    
    # 1. Openpyxl로 엑셀 열기 (수식이 아닌 결과값만 가져오기 위해 data_only=True)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    # 2. LLM에게 보낼 '그리드 데이터' 샘플링
    # 너무 많은 데이터를 보내면 토큰 낭비이므로, 상위 50행 / 최대 10열(J열)까지만 추출
    sample_grid = []
    
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
        row_data = {}
        has_value = False
        
        for cell in row:
            # None이거나 공백인 경우 빈 문자열("")로 처리
            val = str(cell.value).strip() if cell.value is not None else ""
            row_data[cell.column_letter] = val
            if val:
                has_value = True
                
        # 행 전체가 완전히 비어있지 않은 경우에만 샘플에 포함 (노이즈 최소화)
        if has_value:
            sample_grid.append({
                "row_number": row[0].row,
                "cells": row_data
            })

    # 추출된 그리드 데이터를 JSON 문자열로 변환
    grid_json_str = json.dumps(sample_grid, ensure_ascii=False)

    # 3. LLM에게 패턴 분석 요청
    system_prompt = """
    너는 비정형 엑셀 데이터 구조 분석 전문가야.
    사용자가 엑셀 파일의 일부(행 번호와 각 열의 데이터)를 JSON 형태로 줄 거야.
    엑셀 상단에는 제목, 안내문, 버튼 등 무의미한 노이즈가 있을 수 있어.
    
    너의 목표는 '실제 질문들이 시작되는 위치'를 찾는 거야.
    규칙:
    1. '질문' 컬럼에는 사용자의 문의 내용이나 질문 텍스트가 들어있어.
    2. '답변' 컬럼은 질문과 **같은 행**에 위치하며, 현재는 답변을 채워넣기 위해 **비어있어("")**.
    3. 노이즈(상단 안내문 등)를 무시하고, 실제 질문과 답변 데이터 패턴이 반복적으로 시작되는 첫 번째 행을 찾아.
    
    반드시 아래 JSON 형식으로만 응답해:
    {
        "question_col": "C",  // 질문이 있는 열 알파벳 (예: A, B, C)
        "answer_col": "D",    // 비어있는 답변 대상 열 알파벳
        "start_row": 5        // 실제 데이터가 시작되는 행 번호 (정수)
    }
    """

    print("LLM에게 그리드 패턴 분석 요청 중...")
    response = openai.chat.completions.create(
        model="gpt-4o", # 복잡한 패턴 인식에는 성능이 좋은 모델 추천
        response_format={ "type": "json_object" },
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": f"다음은 엑셀 그리드 데이터야:\n\n{grid_json_str}"}
        ],
        temperature=0.0 # 환각 방지를 위해 온도 0
    )
    
    # 4. 결과 반환
    result_json = response.choices[0].message.content
    coordinates = json.loads(result_json)
    
    print(f"✅ 좌표 탐색 완료! 질문 열: {coordinates['question_col']}, 답변 열: {coordinates['answer_col']}, 시작 행: {coordinates['start_row']}")
    
    return coordinates

# --- 실행 예시 ---
# coords = find_qna_coordinates("customer_questions.xlsx", "Sheet1")


import openpyxl
import json
import openai
from openpyxl.utils import get_column_letter  # ⭐️ 추가해야 할 모듈

# ... [이전 코드 생략] ...

def find_qna_coordinates(file_path, sheet_name):
    print(f"[{sheet_name}] 시트 좌표 탐색 시작...")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    sample_grid = []
    
    for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=10):
        row_data = {}
        has_value = False
        current_row_number = None
        
        # enumerate를 사용해 1번부터 시작하는 컬럼 인덱스를 직접 가져옵니다.
        for col_idx, cell in enumerate(row, start=1):
            
            # 1. 컬럼 알파벳 안전하게 추출 (MergedCell 에러 원천 차단)
            col_letter = get_column_letter(col_idx)
            
            # 행 번호 기록 (첫 번째 셀에서 한 번만 가져옵니다)
            if current_row_number is None:
                current_row_number = getattr(cell, 'row', row[0].row)
            
            # 2. 값(Value) 안전하게 추출
            # MergedCell은 value 속성이 아예 없거나 None일 수 있으므로 getattr() 사용
            raw_val = getattr(cell, 'value', None)
            val = str(raw_val).strip() if raw_val is not None else ""
            
            row_data[col_letter] = val
            if val:
                has_value = True
                
        # 행 전체가 비어있지 않은 경우에만 샘플링 추가
        if has_value and current_row_number is not None:
            sample_grid.append({
                "row_number": current_row_number,
                "cells": row_data
            })

    # 추출된 그리드 데이터를 JSON 문자열로 변환
    grid_json_str = json.dumps(sample_grid, ensure_ascii=False)

    # ... [이후 LLM 호출 프롬프트 코드는 동일하게 유지] ...
